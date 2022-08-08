import json
import jsonpath
import requests
import openpyxl
from openpyxl import Workbook

def test_and_multiply_students():
    # API code
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    fileToOpen = open("C:/Users/pavel/PycharmProjects/Python_QA_automation_training/dataDriven/Add_Student.json")
    json_request = json.loads(fileToOpen.read())

    # Open workbook - excel code
    vk = openpyxl.load_workbook('C:/Users/pavel/PycharmProjects/Python_QA_automation_training/dataDriven/Add_student.xlsx')
    sh = vk('Sheet1')

    # How many rows are in table
    rows = sh.max_row

    # Starting loop from second row, as first is header
    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=1, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['last_name'] = cell_date_of_birth.value

        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201